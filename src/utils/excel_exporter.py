"""
Excel Exporter - Export paper data to Excel format
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict
import json
import logging
import os
from pathlib import Path

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel Exporter, supports formatting and beautification"""

    def __init__(self):
        """Initialize exporter"""
        self.default_output_dir = "exports/excel"

    def export(
        self,
        papers: List[Dict],
        output_path: str,
        include_abstract: bool = False,
        sort_by: str = "priority"
    ) -> str:
        """
        Export papers to Excel

        Args:
            papers: List of paper data (including priority, matched_keywords, etc.)
            output_path: Output file path
            include_abstract: Whether to include abstract column
            sort_by: Sort field ('priority', 'citations', 'year')

        Returns:
            Actual saved file path
        """
        if not papers:
            logger.warning("⚠️  No paper data to export")
            return None

        logger.info(f"📊 Preparing to export {len(papers)} papers to Excel...")

        # Ensure output directory exists
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        # Prepare data
        data = []
        for paper in papers:
            # Parse JSON fields
            authors = self._parse_json_field(paper.get('authors', '[]'))
            matched_kws = self._parse_json_field(paper.get('matched_keywords', '[]'))
            fields = self._parse_json_field(paper.get('fields_of_study', '[]'))

            row = {
                'Title': paper.get('title', 'Unknown'),
                'First Author': paper.get('first_author', 'Unknown'),
                'Year': paper.get('year', 'N/A'),
                'Matched Keywords': ', '.join(matched_kws) if matched_kws else 'None',
                'Priority': paper.get('priority', 3),
                'Citations': paper.get('citation_count', 0),
                'Venue': paper.get('venue', 'N/A'),
                'Fields': ', '.join(fields[:3]) if fields else 'N/A',  # Only show first 3 fields
                'DOI': paper.get('doi', 'N/A'),
                'URL': paper.get('url', 'N/A')
            }

            # Optional: add abstract
            if include_abstract:
                row['Abstract'] = paper.get('abstract', 'No abstract')[:500]  # Limit length

            # Optional: add analysis reason
            if 'analysis_reason' in paper:
                row['Analysis Reason'] = paper.get('analysis_reason', '')

            data.append(row)

        # Create DataFrame
        df = pd.DataFrame(data)

        # Sort
        if sort_by == 'priority':
            df = df.sort_values(
                by=['Priority', 'Citations'],
                ascending=[False, False]
            )
        elif sort_by == 'citations':
            df = df.sort_values(by='Citations', ascending=False)
        elif sort_by == 'year':
            df = df.sort_values(by='Year', ascending=False)

        # Export to Excel
        df.to_excel(output_path, index=False, engine='openpyxl')

        # Beautify Excel
        self._format_excel(output_path)

        logger.info(f"✅ Excel export successful: {output_path}")
        return output_path

    def _parse_json_field(self, field):
        """Parse JSON field"""
        if isinstance(field, str):
            try:
                return json.loads(field)
            except:
                return []
        elif isinstance(field, list):
            return field
        else:
            return []

    def _format_excel(self, file_path: str):
        """Beautify Excel format"""
        try:
            wb = load_workbook(file_path)
            ws = wb.active

            # Define styles
            header_fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal='center', vertical='center')

            # Border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Format header
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Priority coloring
            priority_colors = {
                5: "C6EFCE",  # Dark green
                4: "FFEB9C",  # Yellow
                3: "FFC7CE",  # Light red
            }

            # Find Priority column index
            priority_col_idx = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Priority':
                    priority_col_idx = idx
                    break

            # Color rows by Priority
            if priority_col_idx:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    priority = row[priority_col_idx - 1].value
                    if priority in priority_colors:
                        fill = PatternFill(
                            start_color=priority_colors[priority],
                            end_color=priority_colors[priority],
                            fill_type="solid"
                        )
                        for cell in row:
                            cell.fill = fill
                            cell.border = thin_border

            # Adjust column width
            column_widths = {
                'A': 60,  # Title
                'B': 20,  # First Author
                'C': 8,   # Year
                'D': 30,  # Matched Keywords
                'E': 10,  # Priority
                'F': 10,  # Citations
                'G': 20,  # Venue
                'H': 30,  # Fields
                'I': 25,  # DOI
                'J': 40,  # URL
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Freeze header
            ws.freeze_panes = 'A2'

            # Add filter
            ws.auto_filter.ref = ws.dimensions

            wb.save(file_path)
            logger.info(f"   ✅ Excel formatting completed")

        except Exception as e:
            logger.warning(f"⚠️  Excel formatting failed: {e}")

    def export_with_relationships(
        self,
        papers: List[Dict],
        relationships: List[Dict],
        output_path: str
    ) -> str:
        """
        Export papers and relationships to Excel (multiple sheets)

        Args:
            papers: List of papers
            relationships: List of relationships
            output_path: Output path

        Returns:
            Saved file path
        """
        logger.info(f"📊 Preparing to export papers and relationships to Excel...")

        # Ensure output directory exists
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        # Create ExcelWriter
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Papers list
            papers_data = []
            for paper in papers:
                authors = self._parse_json_field(paper.get('authors', '[]'))
                matched_kws = self._parse_json_field(paper.get('matched_keywords', '[]'))

                papers_data.append({
                    'Paper ID': paper.get('paper_id', ''),
                    'Title': paper.get('title', ''),
                    'First Author': paper.get('first_author', ''),
                    'Year': paper.get('year', ''),
                    'Priority': paper.get('priority', 3),
                    'Citations': paper.get('citation_count', 0),
                    'Matched Keywords': ', '.join(matched_kws),
                    'DOI': paper.get('doi', '')
                })

            df_papers = pd.DataFrame(papers_data)
            df_papers.to_excel(writer, sheet_name='Papers', index=False)

            # Sheet 2: Relationships list
            if relationships:
                rel_data = []
                paper_dict = {p.get('paper_id'): p for p in papers}

                for rel in relationships:
                    source_id = rel.get('source_paper_id', '')
                    target_id = rel.get('target_paper_id', '')

                    source_paper = paper_dict.get(source_id, {})
                    target_paper = paper_dict.get(target_id, {})

                    rel_data.append({
                        'Source Paper': source_paper.get('title', source_id)[:50],
                        'Relationship Type': rel.get('relationship_type', 'cites'),
                        'Target Paper': target_paper.get('title', target_id)[:50],
                        'Description': rel.get('relationship_desc', ''),
                        'Source Year': source_paper.get('year', ''),
                        'Target Year': target_paper.get('year', '')
                    })

                df_rel = pd.DataFrame(rel_data)
                df_rel.to_excel(writer, sheet_name='Relationships', index=False)

        logger.info(f"✅ Multi-sheet Excel export successful: {output_path}")
        return output_path


if __name__ == "__main__":
    # Test Excel exporter
    print("🍃 Testing Excel Exporter\n")

    exporter = ExcelExporter()

    # Create test data
    test_papers = [
        {
            'paper_id': 'paper1',
            'title': 'Deep Learning for Autonomous Driving',
            'authors': '["John Smith", "Jane Doe"]',
            'first_author': 'John Smith',
            'year': 2020,
            'priority': 5,
            'citation_count': 150,
            'matched_keywords': '["autonomous driving", "deep learning"]',
            'analysis_reason': 'Highly relevant, directly discusses deep learning methods for autonomous driving',
            'venue': 'CVPR 2020',
            'doi': '10.1109/CVPR.2020.12345',
            'url': 'https://example.com/paper1',
            'fields_of_study': '["Computer Science", "AI"]',
            'abstract': 'This paper presents a novel deep learning approach...'
        },
        {
            'paper_id': 'paper2',
            'title': 'Machine Learning in Robotics',
            'authors': '["Alice Wang"]',
            'first_author': 'Alice Wang',
            'year': 2021,
            'priority': 4,
            'citation_count': 80,
            'matched_keywords': '["machine learning"]',
            'analysis_reason': 'Relevant, discusses machine learning applications in robotics',
            'venue': 'ICRA 2021',
            'doi': '10.1109/ICRA.2021.54321',
            'url': 'https://example.com/paper2',
            'fields_of_study': '["Robotics", "AI"]',
            'abstract': 'We apply machine learning techniques...'
        },
        {
            'paper_id': 'paper3',
            'title': 'Computer Vision Basics',
            'authors': '["Bob Johnson"]',
            'first_author': 'Bob Johnson',
            'year': 2019,
            'priority': 3,
            'citation_count': 200,
            'matched_keywords': '[]',
            'analysis_reason': 'Partially relevant, computer vision fundamentals',
            'venue': 'ICCV 2019',
            'doi': '10.1109/ICCV.2019.11111',
            'url': 'https://example.com/paper3',
            'fields_of_study': '["Computer Vision"]',
            'abstract': 'An introduction to computer vision...'
        }
    ]

    # Test 1: Basic export
    print("=" * 60)
    print("Test 1: Basic Excel Export")
    print("=" * 60)

    output_file = "test_papers_export.xlsx"
    result = exporter.export(
        papers=test_papers,
        output_path=output_file,
        include_abstract=False,
        sort_by='priority'
    )

    if result and os.path.exists(result):
        file_size = os.path.getsize(result) / 1024
        print(f"✅ Export successful: {result}")
        print(f"   File size: {file_size:.1f} KB")
        print(f"   Number of papers: {len(test_papers)} papers")

    # Test 2: Include abstract
    print("\n" + "=" * 60)
    print("Test 2: Export with abstract")
    print("=" * 60)

    output_file2 = "test_papers_with_abstract.xlsx"
    result2 = exporter.export(
        papers=test_papers,
        output_path=output_file2,
        include_abstract=True,
        sort_by='citations'
    )

    if result2:
        print(f"✅ Export successful: {result2}")

    # Test 3: Multi-sheet export
    print("\n" + "=" * 60)
    print("Test 3: Multi-sheet export (with relationships)")
    print("=" * 60)

    test_relationships = [
        {
            'source_paper_id': 'paper1',
            'target_paper_id': 'paper3',
            'relationship_type': 'builds_on',
            'relationship_desc': 'A builds on B computer vision theory'
        }
    ]

    output_file3 = "test_papers_with_relationships.xlsx"
    result3 = exporter.export_with_relationships(
        papers=test_papers,
        relationships=test_relationships,
        output_path=output_file3
    )

    if result3:
        print(f"✅ Multi-sheet export successful: {result3}")

    print("\n🎉 All tests completed!")
    print(f"\nGenerated files:")
    for f in [output_file, output_file2, output_file3]:
        if os.path.exists(f):
            print(f"   - {f}")
